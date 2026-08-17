import os
import io
import re
import zipfile
from PIL import Image
import openpyxl
import pandas as pd
import streamlit as st
import google.generativeai as genai

st.set_page_config(page_title="سامانه ارزیابی هوشمند بازرسی پل‌ها", layout="wide")
st.title("🏗️ سامانه تطبیق چک‌لیست و ارزیابی هوشمند تصاویر پل‌ها")

# تابع تبدیل فایل اکسل چک‌لیست به متن جدول‌بندی‌شده
def excel_to_text(file_or_path):
    try:
        excel_data = pd.read_excel(file_or_path, sheet_name=None)
        text_output = ""
        for sheet_name, df in excel_data.items():
            df_cleaned = df.dropna(how='all').dropna(axis=1, how='all')
            text_output += f"\n--- شیت: {sheet_name} ---\n"
            text_output += df_cleaned.to_markdown(index=False) + "\n"
        return text_output
    except Exception as e:
        return f"خطا در خواندن فایل اکسل: {str(e)}"

# بهینه‌سازی و فشرده‌سازی تصاویر برای ارسال به API
def prepare_image_from_bytes(image_bytes, max_dim=800):
    try:
        with Image.open(io.BytesIO(image_bytes)) as img:
            img = img.convert("RGB")
            if max(img.size) > max_dim:
                img.thumbnail((max_dim, max_dim), Image.Resampling.LANCZOS)
            buffer = io.BytesIO()
            img.save(buffer, format="JPEG", quality=70, optimize=True)
            return {
                "mime_type": "image/jpeg",
                "data": buffer.getvalue()
            }
    except Exception:
        return None

# استخراج ساختاریافته راهنمای بازرسی بر اساس قالب دقیق HELP S و HELP C
def extract_structured_guide(guide_file_or_path, bridge_type="فلزی (HELP S)"):
    sheet_name = "HELP S" if "فلزی" in bridge_type else "HELP C"
    prompt_elements = []
    
    try:
        if isinstance(guide_file_or_path, str):
            with open(guide_file_or_path, 'rb') as f:
                wb = openpyxl.load_workbook(f, data_only=True)
        else:
            wb = openpyxl.load_workbook(guide_file_or_path, data_only=True)
            
        if sheet_name not in wb.sheetnames:
            sheet_name = wb.sheetnames[0]
            
        ws = wb[sheet_name]
        
        # استخراج لنگر و ردیف تصاویر
        row_images = {}
        for img in getattr(ws, '_images', []):
            r = None
            if hasattr(img.anchor, '_from'):
                r = img.anchor._from.row + 1
            elif isinstance(img.anchor, str):
                m = re.search(r'\d+', img.anchor)
                if m: r = int(m.group())
            if r:
                try:
                    row_images.setdefault(r, []).append(img._data())
                except Exception:
                    pass

        current_element = ""
        current_damage = ""
        
        prompt_elements.append(f"=== راهنمای مرجع ارزیابی و تطبیق تصاویر ({sheet_name}) ===")
        
        # خواندن سطر به سطر با لحاظ کردن سلول‌های ادغام‌شده
        for r in range(5, min(ws.max_row + 1, 350)):
            c_elem = ws.cell(row=r, column=17).value
            c_dmg = ws.cell(row=r, column=11).value
            c_sev = ws.cell(row=r, column=6).value
            c_desc = ws.cell(row=r, column=5).value
            
            if c_elem is not None and str(c_elem).strip():
                current_element = str(c_elem).strip()
            if c_dmg is not None and str(c_dmg).strip():
                current_damage = str(c_dmg).strip()
                
            imgs = row_images.get(r, [])
            
            if c_sev or c_desc or imgs:
                desc_text = str(c_desc).strip() if c_desc else "-"
                sev_text = str(c_sev).strip() if c_sev else "-"
                
                header = f"\n[المان: {current_element} | نوع آسیب: {current_damage} | سطح شدت: {sev_text}]\nمعیار ارزیابی: {desc_text}"
                prompt_elements.append(header)
                
                for img_data in imgs:
                    prep = prepare_image_from_bytes(img_data, max_dim=600)
                    if prep:
                        prompt_elements.append(prep)
                        
    except Exception as e:
        prompt_elements.append(f"خطا در استخراج راهنما: {str(e)}")
        
    return prompt_elements

# استخراج چند نمونه شاخص از فایل کامنت‌ها جهت هدایت لحن و اصول ارزیابی
def load_sample_comments(file_or_path, max_samples=15):
    try:
        df = pd.read_excel(file_or_path)
        samples = []
        for _, row in df.dropna().head(max_samples).iterrows():
            samples.append(f"- پل {row.iloc[0]}:\n  «{row.iloc[1]}»")
        return "\n\n".join(samples)
    except Exception:
        return ""

# نوار کناری تنظیمات
st.sidebar.header("⚙️ تنظیمات و مراجع")
api_key = st.sidebar.text_input("کلید Gemini API:", type="password")
bridge_type = st.sidebar.selectbox("نوع سازه پل:", ["فلزی (HELP S)", "بتنی (HELP C)"])

if api_key:
    genai.configure(api_key=api_key, transport='rest')

    guide_path = os.path.join("guides", "inspection_guide.xlsx")
    samples_path = os.path.join("guides", "sample_comments.xlsx")

    st.header("بارگذاری اطلاعات و تصاویر پل")
    col1, col2 = st.columns(2)
    with col1:
        checklist_file = st.file_uploader("۱. فایل چک‌لیست مشاور (اکسل):", type=['xlsx', 'xls'])
    with col2:
        uploaded_zip = st.file_uploader("۲. فایل فشرده عکس‌های پل (ZIP):", type=['zip'])

    if checklist_file:
        with st.expander("👁️ پیش‌نمایش چک‌لیست مشاور"):
            st.dataframe(pd.read_excel(checklist_file))

    images_dict = {}
    if uploaded_zip:
        try:
            with zipfile.ZipFile(uploaded_zip) as z:
                for filename in z.namelist():
                    if filename.lower().endswith(('.jpg', '.jpeg', '.png')) and not filename.startswith('__MACOSX'):
                        base_name = os.path.basename(filename)
                        if base_name:
                            images_dict[base_name] = z.read(filename)
            st.success(f"تعداد {len(images_dict)} تصویر از فایل زیپ استخراج شد.")
        except Exception as e:
            st.error(f"خطا در خواندن فایل زیپ: {e}")

    if checklist_file and images_dict:
        if st.button("🚀 شروع ارزیابی هوشمند"):
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            try:
                status_text.text("در حال بارگذاری تصاویر و ضوابط شیت راهنمای بازرسی...")
                guide_elements = []
                if os.path.exists(guide_path):
                    guide_elements = extract_structured_guide(guide_path, bridge_type)
                
                samples_text = ""
                if os.path.exists(samples_path):
                    samples_text = load_sample_comments(samples_path)

                model = genai.GenerativeModel(model_name="gemini-1.5-pro")
                checklist_content = excel_to_text(checklist_file)
                
                batch_size = 6
                batch_analyses = []
                image_items = list(images_dict.items())
                total_batches = (len(image_items) + batch_size - 1) // batch_size
                
                # مرحله ۱: تحلیل دسته‌ای تصاویر با انطباق بر تصاویر راهنما
                for i in range(0, len(image_items), batch_size):
                    batch_num = (i // batch_size) + 1
                    current_batch = image_items[i:i + batch_size]
                    status_text.text(f"در حال تحلیل دسته تصاویر {batch_num} از {total_batches}...")
                    
                    images_data = [prepare_image_from_bytes(data) for name, data in current_batch]
                    images_data = [img for img in images_data if img is not None]
                    batch_names = [name for name, data in current_batch]
                    
                    prompt_parts = [
                        "شما ارزیاب ارشد سازه و بازرس تخصصی پل‌ها هستید.",
                        "الگوها، تعاریف و تصاویر مرجع راهنمای بازرسی:",
                    ] + guide_elements + [
                        f"\nاکنون تصاویر میدانی ارسالی را تحلیل کن (فایل‌ها: {', '.join(batch_names)}):",
                        "۱. المان‌ها (تیر اصلی، تیر فرعی/دال، ستون، نرده، کف‌پله، فونداسیون، تابلو تبلیغاتی و تأسیسات) را شناسایی کن.",
                        "۲. عیوب را با تصاویر و معیارهای راهنما تطبیق بده و شدت (اضطراری/متوسط/کم) را تعیین کن.",
                        "۳. میان آسیب‌های سازه‌ای واقعی با عیوب اجرایی اولیه (مثل کج‌سلیقگی اجرایی ورق‌ها، سوراخکاری زهکشی یا آثار حین نصب) تمایز قائل شو."
                    ] + images_data

                    res = model.generate_content(prompt_parts)
                    batch_analyses.append(f"تحلیل تصاویر ({', '.join(batch_names)}):\n{res.text}")
                    progress_bar.progress((batch_num) / (total_batches + 1))

                # مرحله ۲: ممیزی چک‌لیست مشاور و تدوین کامنت نهایی
                status_text.text("در حال ممیزی چک‌لیست مشاور و تدوین کامنت نهایی...")
                all_visual_data = "\n\n".join(batch_analyses)
                
                final_prompt = f"""
                مشاهدات استخراج‌شده از تصاویر میدانی پل:
                ===================
                {all_visual_data}
                ===================

                چک‌لیست تکمیل‌شده توسط مشاور:
                ===================
                {checklist_content}
                ===================

                الگوی نگارش، لحن ممیزی و نمونه کامنت‌های مورد تایید کارفرما:
                ===================
                {samples_text}
                ===================

                دستور کار ممیزی نهایی:
                ۱. تطبیق دقیق ادعاهای چک‌لیست مشاور با مشاهدات عکس‌ها (با ذکر مستقیم شماره عکس‌ها مانند «در عکس شماره...»).
                ۲. بیان مغایرت‌ها (ثبت نادرست المان، کم‌نمایی یا اغراق در شدت خرابی، نقص ثبت‌نشده مانند مفقودی پیچ فلنج‌ها یا سرقت اعضای قائم نرده، و بررسی وضعیت تابلو تبلیغاتی/تأسیسات عبوری/فونداسیون).
                ۳. نگارش کامنت نهایی ممیزی، یکپارچه، فنی، صریح و کاملاً منطبق بر سبک نمونه کامنت‌های کارفرما.
                """
                
                final_response = model.generate_content(final_prompt)
                progress_bar.progress(1.0)
                status_text.empty()
                progress_bar.empty()
                
                st.subheader("📋 نتیجه ارزیابی و کامنت ممیزی پیشنهادی:")
                st.markdown(final_response.text)
                
            except Exception as ex:
                status_text.empty()
                progress_bar.empty()
                st.error(f"خطا در پردازش: {str(ex)}")
else:
    st.warning("لطفاً کلید API را وارد کنید.")